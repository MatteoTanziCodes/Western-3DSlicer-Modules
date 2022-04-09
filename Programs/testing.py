from vpython import *
from openpyxl import *
import random
import numpy as np
import os

##### Testing file #####

#### Main line segment ####
# Create main line segment vector
vector_array_main = [[0,0,0]]
branch_heads_array = []
branch_array = []
MAIN_LINE_LENGTH = 80
BRANCH_AMOUNT_DIVISOR = 4
BRANCH_AMOUNT_UPPER = MAIN_LINE_LENGTH/BRANCH_AMOUNT_DIVISOR
AMOUNT_OF_BRANCHES = 0
RANDOM_NUM_UPPER = 20
RANDOM_NUM_LOWER = 1
RANDOM_NUM_DIVISOR = 10
X_TRANSLATION = 20
X_STRETCH = 1/4
Y_TRANSLATION = 0
Y_STRETCH = 0
Z_TRANSLATION = 0
Z_STRETCH = 1.5


def next_vector(vector_array, direction, index):
    prev_head = vector_array[index]

    new_x = random_number_increase(abs(prev_head[0]))
    new_y = random_number_increase(abs(prev_head[1]))
    new_z = random_number_increase(abs(prev_head[2]))

    if direction == "x":
        new_head = [round(prev_head[0],1),round(new_y,1),round(new_z,1)]
    elif direction == "y":
        new_head = [round(new_x,1),round(prev_head[1],1),round(new_z,1)]
    elif direction == "-y":
        new_head = [-1 * round(new_x,1),round(prev_head[1],1),round(new_z,1)]
    elif direction == "z":
        new_head = [round(new_x,1),round(new_y,1),round(prev_head[2],1)]
    elif direction == "-z":
        new_head = [-1 * round(new_x,1),round(new_y,1),round(prev_head[2],1)]
    return new_head

def random_number_increase(num):
    random_num = random.randint(RANDOM_NUM_LOWER,RANDOM_NUM_UPPER)
    random_num = random_num/RANDOM_NUM_DIVISOR
    num = num + random_num
    return num

def create_main_line(vector_array_main):
    for x in range(MAIN_LINE_LENGTH):
        vector_array_main.append(next_vector(vector_array_main, "x", x))

def display_main_line(vector_array_main, color_choice):
    main_line = curve(vector_array_main[0],vector_array_main[1], color = color_choice, radius = 0.5)
    main_line.append(vector_array_main)

def pick_branch_nodes(vector_array_main, branch_heads_array):
    AMOUNT_OF_BRANCHES = random.randint(BRANCH_AMOUNT_UPPER - 8,BRANCH_AMOUNT_UPPER)
    index_on_main_line = random.randint(2,BRANCH_AMOUNT_DIVISOR)
    for branch in range(AMOUNT_OF_BRANCHES):
        branch_heads_array.append(vector_array_main[index_on_main_line])
        index_on_main_line = index_on_main_line + random.randint(2,BRANCH_AMOUNT_DIVISOR)
    return branch_heads_array

def create_branches(branch_heads_array):
    amount_of_branches = len(branch_heads_array)
    branch_array = []
    for x in range(amount_of_branches):
        branch_array.append([])
        branch_array[x].append(branch_heads_array[x])
        direction = random.randint(1,4)
        if direction == 1:
            vector_direction = "y"
        elif direction == 2:
            vector_direction = "z"
        elif direction == 3:
            vector_direction = "-y"
        else:
            vector_direction = "-z"
        BRANCH_LENGTH = random.randint(8,25)
        for y in range(BRANCH_LENGTH):
            branch_array[x].append(next_vector(branch_array[x], vector_direction, y))
    return branch_array
            
def display_branches(branch_array, color_choice):
    array_of_branch_curves = []
    for x in range(len(branch_array)):
        array_of_branch_curves.append(curve(branch_array[x][0], branch_array[x][1], color = color_choice, radius = 0.3))
        array_of_branch_curves[x].append(branch_array[x])

def create_target():
    random_x = random.randint(5,40)
    random_y = random.randint(5,40)
    random_z = random.randint(5,40)
    target = (random_x, random_y, random_z)
    return target

def bifurcation_points(branch_heads_array):
    return branch_heads_array

def centerline_points(vector_array_main):
    centerline_points_array = []
    index_on_main_line = random.randint(2,BRANCH_AMOUNT_DIVISOR)
    for branch in range(AMOUNT_OF_BRANCHES):
        centerline_points_array.append(vector_array_main[index_on_main_line])
        index_on_main_line = index_on_main_line + random.randint(2,BRANCH_AMOUNT_DIVISOR)
    return centerline_points_array

def point_transformation(vector):
    new_vector = []
    new_vector.append((vector[0] + X_TRANSLATION) * X_STRETCH)
    new_vector.append((vector[1] + Y_TRANSLATION) * Y_STRETCH)
    new_vector.append((vector[2] + Z_TRANSLATION) * Z_STRETCH)
    return new_vector

def deform_main_vessel(vector_array_main):
    deformed_main_array = []
    for x in range(len(vector_array_main)):
        deformed_main_array.append(point_transformation(vector_array_main[x]))
    return deformed_main_array

def deform_branches(branch_array):
    deformed_branch_array = []
    for x in range(len(branch_array)):
        deformed_branch_array.append([])
        deformed_branch_array[x].append(point_transformation(branch_array[x][0]))
        for y in range(len(branch_array[x])):
            deformed_branch_array[x].append(point_transformation(branch_array[x][y]))
    return deformed_branch_array

def measure_FRE(CT_fiducials, US_fiducials):
    squared_dist_array = []
    for x in range(len(CT_fiducials)):
        p1 = np.array([CT_fiducials[x][0], CT_fiducials[x][1], CT_fiducials[x][2]])
        p2 = np.array([US_fiducials[x][0], US_fiducials[x][1], US_fiducials[x][2]])
        squared_dist = np.sum((p1-p2)**2, axis=0)
        squared_dist_array.append(squared_dist)
    return squared_dist_array

def mearure_TRE(CT_Target, US_Target):
    p1 = np.array([CT_Target[0], CT_Target[1], CT_Target[2]])
    p2 = np.array([US_Target[0], US_Target[1], US_Target[2]])
    squared_dist = np.sum((p1-p2)**2, axis=0)
    return squared_dist

# def ICP_transformation(CT_fiducials, US_fiducials):
# def CPD_transformation(CT_fiducials, US_fiducials):

def create_file(page_names, file_name):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = page_names[0]
    ws2 = wb.create_sheet(page_names[1])
    ws3 = wb.create_sheet(page_names[2])
    ws4 = wb.create_sheet(page_names[3])
    wb.save(filename = file_name)

def delete_file(file_name):
    if os.path.exists(file_name + ".xlsx"):
        os.remove(file_name + ".xlsx")
    else:
        print("The file does not exist")

def append_data(page_name,file_name, FREs, TRE):
    wb = load_workbook(file_name + ".xlsx")
    wb.active = wb[page_name]
    ws = wb.active
    # append first empty row with FREs and TRE

def data_measurement(page_name,file_name):
    measurements = []
    FREs = []
    TREs = []
    fre_sum = 0
    tre_sum = 0
    wb = load_workbook(file_name + ".xlsx")
    wb.active = wb[page_name]
    ws = wb.active

    # Read all FREs and TREs
    

    # calculate means
    for x in range(len(FREs)):
        fre_sum = fre_sum + FREs[x]
    
    for x in range(len(TREs)):
        tre_sum = tre_sum + TREs[x]
    
    measurements.append()
    measurements.append()

    return measurements


create_main_line(vector_array_main)
display_main_line(vector_array_main, color.red)
pick_branch_nodes(vector_array_main, branch_heads_array)
branch_array = create_branches(branch_heads_array)
display_branches(branch_array, color.red)

deformed_main_array = deform_main_vessel(vector_array_main)
deformed_branch_array = deform_branches(branch_array)

display_main_line(deformed_main_array, color.blue)
display_branches(deformed_branch_array, color.blue)

CT_bifurcation_points_array = bifurcation_points(branch_heads_array)
CT_centerline_points_array = centerline_points(vector_array_main)

US_bifurcation_points_array = bifurcation_points(deform_main_vessel(branch_heads_array))
US_centerline_points_array = centerline_points(deformed_main_array)

CT_Target = create_target()
US_Target = point_transformation(CT_Target)

while True:
    pass
